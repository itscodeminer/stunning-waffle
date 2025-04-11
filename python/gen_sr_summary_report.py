import pandas as pd
from sqlalchemy import create_engine
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
import re
import os

# === CONFIG ===
DB_URI = "postgresql+psycopg2://user:password@host:5432/dbname"
REPORT_DIR = "sr_reports"
os.makedirs(REPORT_DIR, exist_ok=True)

# === 1. Connect and Load Full Data ===
engine = create_engine(DB_URI)

with engine.connect() as conn:
    df = pd.read_sql("SELECT * FROM get_severity_summary();", conn)

# === 2. Build breakdown like "2 P1, 3 P5+" ===
df['breakdown'] = df['count'].astype(str) + ' P' + df['bucket']

# === 3. Unique managers
assoc_managers = df['associate_manager'].dropna().unique()
managers = df['manager'].dropna().unique()

# === 4. Reporting Function ===
def generate_excel_report(filtered_df, report_name):
    # Pivot breakdowns
    pivot_df = (
        filtered_df.groupby(['tech_id', 'first_name', 'state', 'city', 'zone', 'team', 'closed_day'])
            .agg({'breakdown': lambda x: ', '.join(sorted(x))})
            .reset_index()
            .pivot(index=['tech_id', 'first_name', 'state', 'city', 'zone', 'team'],
                   columns='closed_day',
                   values='breakdown')
            .reset_index()
    )

    # Rename date columns to "Apr-10" format
    pivot_df.columns = [
        col.strftime('%b-%d') if isinstance(col, pd.Timestamp) else col
        for col in pivot_df.columns
    ]

    # Add Total SRs
    totals = filtered_df.groupby(['tech_id', 'first_name', 'state', 'city', 'zone', 'team'])['count'].sum().reset_index(name='Total_SRs')
    final_df = pd.merge(pivot_df, totals, on=['tech_id', 'first_name', 'state', 'city', 'zone', 'team'])
    cols = [col for col in final_df.columns if col != 'Total_SRs'] + ['Total_SRs']
    final_df = final_df[cols]

    # Save to Excel
    filename = os.path.join(REPORT_DIR, f"{report_name}_SR_Report.xlsx")
    sheet_name = "SR Report"
    final_df.to_excel(filename, index=False, sheet_name=sheet_name)

    # Apply Formatting
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    # Add Excel table formatting
    end_col = get_column_letter(final_df.shape[1])
    table_range = f"A1:{end_col}{final_df.shape[0] + 1}"
    table = Table(displayName="SRReportTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Wrap + Set width for date columns
    wrap_alignment = Alignment(wrap_text=True)
    for idx, column_cells in enumerate(ws.columns, start=1):
        header_value = column_cells[0].value
        col_letter = get_column_letter(idx)

        if isinstance(header_value, str) and re.match(r"^[A-Z][a-z]{2}-\d{2}$", header_value):
            ws.column_dimensions[col_letter].width = 25
            for cell in column_cells:
                cell.alignment = wrap_alignment
        else:
            # Auto width for non-date columns
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)
    return filename

# === 5. Loop Through Managers and Generate Reports ===
report_files = []

for am in assoc_managers:
    team_df = df[df['associate_manager'] == am]
    if not team_df.empty:
        file = generate_excel_report(team_df, f"AM_{am.replace(' ', '_')}")
        report_files.append((am, file, 'Associate Manager'))

for mgr in managers:
    team_df = df[df['manager'] == mgr]
    if not team_df.empty:
        file = generate_excel_report(team_df, f"MGR_{mgr.replace(' ', '_')}")
        report_files.append((mgr, file, 'Manager'))

print("✅ Reports generated:")
for name, file, role in report_files:
    print(f"{role}: {name} → {file}")
