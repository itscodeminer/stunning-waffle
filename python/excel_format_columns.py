import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load your data into a DataFrame
df = pd.read_excel('your_file.xlsx')  # Replace with your actual file path

# Define the column pairs you want to compare
column_pairs = [
    ('dev_datatype', 'qa_datatype'),
    ('prod_datatype', 'qa_datatype'),
    ('staging_datatype', 'qa_datatype'),
    ('dev_datatype', 'prod_datatype'),
    ('staging_datatype', 'prod_datatype')
]

# Use pandas ExcelWriter with openpyxl engine to write the DataFrame to Excel
with pd.ExcelWriter('highlighted_file.xlsx', engine='openpyxl') as writer:
    # Write the DataFrame to the Excel file
    df.to_excel(writer, sheet_name='Sheet1', index=False)

    # Access the openpyxl workbook and worksheet
    workbook  = writer.book
    worksheet = workbook.active
    
    # Define the red fill for highlighting differences
    highlight_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Loop through each pair of columns
    for col1, col2 in column_pairs:
        # Get column indices for openpyxl (1-based indexing)
        col1_idx = df.columns.get_loc(col1) + 1
        col2_idx = df.columns.get_loc(col2) + 1
        
        # Iterate through each row and apply formatting directly
        for row_idx in range(2, len(df) + 2):  # Start from row 2 to skip header row
            # Compare the values and apply formatting if they differ
            if df.at[row_idx - 2, col1] != df.at[row_idx - 2, col2]:
                worksheet.cell(row=row_idx, column=col1_idx).fill = highlight_fill
                worksheet.cell(row=row_idx, column=col2_idx).fill = highlight_fill

# The highlighted file is saved as 'highlighted_file.xlsx'
