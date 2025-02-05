import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine, text
from sqlalchemy.exc import SQLAlchemyError

file_name=r'path\to\your\file'
    
# Specify schema names for Dev and QA
source_schema_name = None  # Replace with actual schema name for Dev
target_schema_name = None  # Replace with actual schema name for QA

source_db_config = {
    'host': None,
    'database': None,
    'user': None,
    'password': None,
    'port': 5432  # Default PostgreSQL port
}

target_db_config = {
    'host': None,
    'database': None,
    'user': None,
    'password': None,
    'port': 5432  # Default PostgreSQL port
}

# Function to create a database connection using SQLAlchemy
def get_engine(db_config):
    try:
        # Format for the connection string: 'postgresql://user:password@host:port/database'
        connection_string = f"postgresql://{db_config['user']}:{db_config['password']}@{db_config['host']}:{db_config['port']}/{db_config['database']}"
        engine = create_engine(connection_string)
        return engine
    except SQLAlchemyError as e:
        print(f"Error in connecting to the database: {e}")
        return None

# Function to fetch the schema of tables from the database using inline SQL
def get_schema(engine, schema_name='public'):
    query = """
                SELECT table_name, column_name, udt_name as data_type, character_maximum_length, is_nullable, column_default
                FROM information_schema.columns
                WHERE table_schema = :schema_name
                ORDER BY table_name, ordinal_position;
            """
    try:
        # Execute the query with the schema name as a parameter
        with engine.connect() as connection:
            result = connection.execute(text(query), {'schema_name': schema_name}).fetchall()
        
        # Organize the results into a dictionary format
        schema = {}
        for row in result:
            table_name = row[0]
            column_name = row[1]
            data_type = row[2]
            max_length = row[3]
            is_nullable = row[4]
            column_default = row[5]
            if table_name not in schema:
                schema[table_name] = []
            schema[table_name].append({
                'column_name': column_name,
                'data_type': data_type,
                'column_default': column_default,
                'max_length': max_length,
                'is_nullable': is_nullable
            })
        return schema
    except SQLAlchemyError as e:
        print(f"Error while fetching schema: {e}")
        return {}

# Compare two schemas (Dev vs QA)
def compare_tables(source_schema, target_schema):
    changes = {'added': [], 'removed': [], 'modified': []}
    
    # Compare added or removed tables
    source_tables = set(source_schema.keys())
    target_tables = set(target_schema.keys())
    
    # Updated logic:
    # - "added" tables are in Dev but not in QA (changes made first in Dev)
    # - "removed" tables are in QA but not in Dev (removed in Dev, but still in QA)
    changes['added'] = list(source_tables - target_tables)
    changes['removed'] = list(target_tables - source_tables)    
                
    return changes

# Generate SQL queries for changes
def compare_columns_and_generate_sql_queries(source_schema, target_schema, source_schema_name='public', target_schema_name='public'):
    sql_queries = []

    # Generate SQL for added, removed, and modified columns
    for table_name, source_columns in source_schema.items():
        if table_name in target_schema:
            target_columns = target_schema[table_name]

            # Compare added and removed columns
            source_column_names = {col['column_name'] for col in source_columns}
            target_column_names = {col['column_name'] for col in target_columns}

            # Added columns (in source but not in target)
            added_columns = source_column_names - target_column_names
            for column_name in added_columns:
                source_column = next(col for col in source_columns if col['column_name'] == column_name)
                data_type = source_column['data_type']
                max_length = source_column['max_length']
                is_nullable = source_column['is_nullable']
                column_default = source_column['column_default']
                add_column_query = f"ALTER TABLE {target_schema_name}.{table_name} ADD COLUMN {column_name} {data_type} "
                if max_length:
                    add_column_query += f"({max_length}) "
                if is_nullable.lower() == "no":
                    add_column_query += "NOT NULL"
                else:
                    add_column_query += " NULL"
                
                if column_default:
                    add_column_query += f" SET DEFAULT {column_default}"
                    
                add_column_query += ";"
                sql_queries.append({
                    'table': table_name,
                    'column': column_name,
                    'change': 'added',
                    'source_data_type': source_column['data_type'],
                    'target_data_type': '',
                    'source_is_nullable': source_column['is_nullable'],
                    'target_is_nullable': '',
                    'source_max_length': source_column['max_length'],
                    'target_max_length': '',
                    'source_default': source_column['column_default'],
                    'target_default': '',
                    'sql_query': add_column_query
                })

            # Removed columns (in target but not in source)
            removed_columns = target_column_names - source_column_names
            for column_name in removed_columns:
                remove_column_query = f"ALTER TABLE {target_schema_name}.{table_name} DROP COLUMN {column_name};"
                sql_queries.append({
                    'table': table_name,
                    'column': column_name,
                    'change': 'removed',
                    'source_data_type': '',
                    'target_data_type': target_column['data_type'],
                    'source_is_nullable': '',
                    'target_is_nullable': target_column['is_nullable'],
                    'source_max_length': '',
                    'target_max_length': target_column['max_length'],
                    'source_default': '',
                    'target_default': target_column['column_default'],
                    'sql_query': remove_column_query
                })

            # Modified columns (compare column attributes)
            for column_name in source_column_names & target_column_names:
                source_column = next(col for col in source_columns if col['column_name'] == column_name)
                target_column = next(col for col in target_columns if col['column_name'] == column_name)

                # If there are differences, generate ALTER COLUMN query
                if (source_column['data_type'] != target_column['data_type'] or
                        source_column['is_nullable'] != target_column['is_nullable'] or
                        source_column['max_length'] != target_column['max_length'] or
                        source_column['column_default'] != target_column['column_default']):
                    modify_column_query = f"ALTER TABLE {target_schema_name}.{table_name} ALTER COLUMN {column_name} "
                    
                    if source_column['data_type'] != target_column['data_type']:
                        modify_column_query += f"SET DATA TYPE {source_column['data_type']} "
                        
                    if source_column['is_nullable'] != target_column['is_nullable']:
                        if source_column['is_nullable'].lower() == 'no':
                            modify_column_query += "SET NOT NULL "
                        else:
                            modify_column_query += "DROP NOT NULL "
                            
                    if source_column['max_length'] != target_column['max_length']:
                        if source_column['max_length']:
                            modify_column_query += f"SET DATA TYPE {source_column['data_type']}({source_column['max_length']}) "
                        else:
                            modify_column_query += f"SET DATA TYPE {source_column['data_type']} "
                    
                    if source_column['column_default'] != target_column['column_default']:
                        if source_column['column_default'] is not None:
                            modify_column_query += f" SET DEFAULT {source_column['column_default']}"
                        else:
                            modify_column_query += " DROP DEFAULT"
                        change_detected = True

                    modify_column_query += ";"
                    sql_queries.append({
                        'table': table_name,
                        'column': column_name,
                        'change': 'modified',
                        'source_data_type': source_column['data_type'],
                        'target_data_type': target_column['data_type'],
                        'source_is_nullable': source_column['is_nullable'],
                        'target_is_nullable': target_column['is_nullable'],
                        'source_max_length': source_column['max_length'],
                        'target_max_length': target_column['max_length'],
                        'source_default': source_column['column_default'],
                        'target_default': target_column['column_default'],
                        'sql_query': modify_column_query
                    })

    return sql_queries

# Write the result to an Excel file
def write_to_excel(table_changes, column_changes):
        
    # Creating dataframes for tables and column changes
    table_changes_df = pd.DataFrame({
        'Change Type': ['Added Tables']*len(table_changes['added']) + ['Removed Tables']*len(table_changes['removed']),
        'Table Name': table_changes['added'] + table_changes['removed'],
        'Details': ['N/A']*len(table_changes['added']) + ['N/A']*len(table_changes['removed'])
    })

    # Add SQL Queries to the column changes DataFrame
    column_changes_df = pd.DataFrame(column_changes)
    
    column_pairs = [
                        ('source_data_type', 'target_data_type'),
                        ('source_is_nullable', 'target_is_nullable'),
                        ('source_max_length', 'target_max_length'),
                        ('source_default', 'target_default')
                   ]

    # Create a Pandas Excel writer using openpyxl engine
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        table_changes_df.to_excel(writer, sheet_name="Table Changes", index=False)
        column_changes_df.to_excel(writer, sheet_name="Column Changes", index=False)        
        
        # Access the openpyxl workbook and worksheet
        workbook  = writer.book
        worksheet = workbook['Column Changes']
        
        # Define the yellow fill for highlighting differences
        highlight_fill = PatternFill(start_color='FCEA04', end_color='FCEA04', fill_type='solid')

        # Loop through each pair of columns
        for col1, col2 in column_pairs:
            # Get column indices for openpyxl (1-based indexing)
            source_col_idx = column_changes_df.columns.get_loc(col1) + 1
            target_col_idx = column_changes_df.columns.get_loc(col2) + 1
            
            # Iterate through each row and apply formatting directly
            for row_idx in range(2, len(column_changes_df) + 2):  # Start from row 2 to skip header row
                # Compare the values and apply formatting if they differ
                if column_changes_df.at[row_idx - 2, col1] != column_changes_df.at[row_idx - 2, col2]:
                    worksheet.cell(row=row_idx, column=source_col_idx).fill = highlight_fill
                    worksheet.cell(row=row_idx, column=target_col_idx).fill = highlight_fill

    print(f"Results written to {file_name}")

# Main function to execute the script
def main():

    # Create engine connections for Dev and QA databases
    source_engine = get_engine(source_db_config)
    target_engine = get_engine(target_db_config)

    if not source_engine or not target_engine:
        print("Failed to connect to one or more databases. Exiting.")
        return
    
    # Fetch schemas with dynamic schema name
    source_schema = get_schema(source_engine, schema_name=source_schema_name)
    target_schema = get_schema(target_engine, schema_name=target_schema_name)
    
    # Generate SQL queries to replicate changes in QA
    column_changes = compare_columns_and_generate_sql_queries(source_schema, target_schema, source_schema_name, target_schema_name)
    
    # Compare schemas
    table_changes = compare_tables(source_schema, target_schema)
    
    # Write the comparison results to an Excel file
    write_to_excel(table_changes, column_changes)

if __name__ == "__main__":
    main()
